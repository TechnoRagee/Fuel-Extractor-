"""
Convert IOCL SQLite Database to Formatted Excel (.xlsx) File
Usage:
    python db_to_excel.py
    python db_to_excel.py --db iocl_outlets.db --output iocl_outlets.xlsx
    python db_to_excel.py --separate-sheets
"""

import os
import sqlite3
import argparse
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def export_db_to_excel(db_path='iocl_outlets.db', output_excel='iocl_outlets.xlsx', separate_sheets=False):
    if not os.path.exists(db_path):
        print(f"[ERROR] Database file '{db_path}' does not exist!")
        return False

    print(f"Connecting to SQLite database: {db_path}...")
    conn = sqlite3.connect(db_path)
    
    # Read entire outlets table
    query = "SELECT * FROM outlets ORDER BY state, city, locality, outlet_name"
    df = pd.read_sql_query(query, conn)
    conn.close()

    total_rows = len(df)
    print(f"Loaded {total_rows} records from database.")
    if total_rows == 0:
        print("[WARNING] Database contains 0 records. Exporting empty template.")

    print(f"Generating formatted Excel file: {output_excel}...")
    
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        # 1. Main Sheet: All Outlets
        df.to_excel(writer, sheet_name='All Outlets', index=False)
        
        # 2. State-wise Breakdown (Summary Sheet)
        if total_rows > 0 and 'state' in df.columns:
            summary_df = df.groupby('state').agg(
                Total_Outlets=('outlet_id', 'count'),
                Cities_Covered=('city', 'nunique'),
                Avg_Rating=('rating_value', lambda x: pd.to_numeric(x, errors='coerce').mean())
            ).reset_index()
            summary_df['Avg_Rating'] = summary_df['Avg_Rating'].round(2)
            summary_df = summary_df.sort_values(by='Total_Outlets', ascending=False)
            summary_df.to_excel(writer, sheet_name='State Summary', index=False)

        # 3. Optional Separate Sheets per State
        if separate_sheets and total_rows > 0 and 'state' in df.columns:
            for state_name, group in df.groupby('state'):
                clean_name = str(state_name)[:30].replace('/', '-').replace('\\', '-')
                if clean_name.strip():
                    group.to_excel(writer, sheet_name=clean_name, index=False)

    # Apply professional styling with openpyxl
    import openpyxl
    wb = openpyxl.load_workbook(output_excel)

    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid') # IOCL Deep Blue
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )

    for sheet in wb.worksheets:
        # Freeze top header row
        sheet.freeze_panes = 'A2'
        sheet.row_dimensions[1].height = 28

        # Style header row
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Auto-fit column widths with padding
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                # Apply light border and alignment
                if cell.row > 1:
                    cell.border = thin_border
                    if cell.value is not None:
                        val_str = str(cell.value)
                        # Check length
                        if len(val_str) > max_len:
                            max_len = len(val_str)
                else:
                    val_str = str(cell.value or '')
                    if len(val_str) > max_len:
                        max_len = len(val_str)
            
            # Set width between 12 and 45
            adjusted_width = min(max(max_len + 3, 12), 45)
            sheet.column_dimensions[col_letter].width = adjusted_width

    wb.save(output_excel)
    print(f"[SUCCESS] Successfully exported {total_rows} records to '{output_excel}'!")
    return True

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Export IOCL SQLite Database to Formatted Excel (.xlsx)')
    parser.add_argument('--db', default='iocl_outlets.db', help='Path to SQLite database file (default: iocl_outlets.db)')
    parser.add_argument('--output', default='iocl_outlets.xlsx', help='Output Excel filename (default: iocl_outlets.xlsx)')
    parser.add_argument('--separate-sheets', action='store_true', help='Create individual sheet for each State')

    args = parser.parse_args()
    export_db_to_excel(db_path=args.db, output_excel=args.output, separate_sheets=args.separate_sheets)
