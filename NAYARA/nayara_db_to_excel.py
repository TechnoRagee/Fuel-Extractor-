"""
Nayara Energy Outlets Database to Excel & CSV Exporter.
Reads nayara_outlets.db and exports a multi-sheet formatted Excel file and CSV.
Features:
- "All Outlets" master worksheet
- "State Summary" summary worksheet with state-wise counts, percentages, and district stats
- Dedicated individual worksheets for top states
- Professional Nayara Green/Orange styling, auto-adjusted column widths, and gridlines
- Exports clean nayara_outlets.csv
"""

import os
import sys
import sqlite3
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_PATH = "nayara_outlets.db"
EXCEL_PATH = "nayara_outlets.xlsx"
CSV_PATH = "nayara_outlets.csv"

# Styling Palette (Nayara Emerald Green / Orange Theme)
HEADER_FILL = PatternFill(start_color="006837", end_color="006837", fill_type="solid")  # Emerald Green
HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Segoe UI", size=14, bold=True, color="006837")
SUBTITLE_FONT = Font(name="Segoe UI", size=10, italic=True, color="555555")
DATA_FONT = Font(name="Segoe UI", size=10)
BOLD_FONT = Font(name="Segoe UI", size=10, bold=True)
SUMMARY_HEADER_FILL = PatternFill(start_color="D35400", end_color="D35400", fill_type="solid") # Deep Orange
ACCENT_FILL = PatternFill(start_color="F2F9F4", end_color="F2F9F4", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
TOP_THICK_BOTTOM_DOUBLE = Border(
    top=Side(style="thin", color="000000"),
    bottom=Side(style="double", color="000000"),
)

DISPLAY_COLUMNS = {
    "cms_code": "CMS Station Code",
    "ro_name": "Station / Dealership Name",
    "address": "Full Address",
    "village": "Village / Locality",
    "taluka": "Taluka / Tehsil",
    "district": "District / City",
    "state": "State / UT",
    "pincode": "Pincode",
    "latitude": "Latitude",
    "longitude": "Longitude",
    "efp": "EFP Status",
    "petrol_price": "Petrol Price (₹/L)",
    "diesel_price": "Diesel Price (₹/L)",
    "created_at": "Extracted At"
}

def style_data_sheet(ws, df, sheet_title):
    """Applies modern professional styling to an outlet data worksheet."""
    ws.views.sheetView[0].showGridLines = True

    # Title Block
    ws.insert_rows(1, 2)
    ws.cell(row=1, column=1, value=f"NAYARA ENERGY — {sheet_title.upper()}").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Total Stations: {len(df):,} | Exported from nayara_outlets.db").font = SUBTITLE_FONT

    header_row = 3

    # Format Headers
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = THIN_BORDER
    ws.row_dimensions[header_row].height = 28

    # Format Data Rows
    for r_idx in range(header_row + 1, header_row + 1 + len(df)):
        ws.row_dimensions[r_idx].height = 20
        is_even = (r_idx % 2 == 0)
        for c_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if is_even:
                cell.fill = ACCENT_FILL

            col_name = df.columns[c_idx - 1]
            if col_name in ["CMS Station Code", "Pincode", "EFP Status"]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name in ["Latitude", "Longitude"]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "0.000000"
            elif "Price" in col_name:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "₹#,##0.00"
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-fit Column Widths
    for col_idx in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=header_row, column=col_idx).value or "")),
            max((len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(header_row + 1, min(header_row + 100, header_row + 1 + len(df)))), default=10)
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

def build_summary_sheet(wb, df):
    """Builds a summary worksheet aggregating counts by State/UT."""
    ws = wb.create_sheet(title="State Summary", index=0)
    ws.views.sheetView[0].showGridLines = True

    # Title Block
    ws.cell(row=1, column=1, value="NAYARA ENERGY — STATE-WISE DISTRIBUTION").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Total Stations Extracted: {len(df):,}").font = SUBTITLE_FONT

    header_row = 4
    headers = ["Sr. No.", "State / Union Territory", "Retail Stations", "% Share", "Districts Covered"]
    for col_idx, h_text in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h_text)
        cell.fill = SUMMARY_HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[header_row].height = 26

    # Group data by State
    state_grp = df.groupby("State / UT").agg(
        Count=("CMS Station Code", "count"),
        Districts=("District / City", "nunique")
    ).reset_index().sort_values(by="Count", ascending=False)

    total_count = len(df)
    current_row = header_row + 1

    for idx, row in state_grp.iterrows():
        sr_no = current_row - header_row
        state_name = row["State / UT"] or "Unspecified"
        count = row["Count"]
        pct = (count / total_count) if total_count > 0 else 0
        districts = row["Districts"]

        ws.cell(row=current_row, column=1, value=sr_no).alignment = Alignment(horizontal="center")
        ws.cell(row=current_row, column=2, value=state_name).alignment = Alignment(horizontal="left")
        c3 = ws.cell(row=current_row, column=3, value=count)
        c3.alignment = Alignment(horizontal="right")
        c3.number_format = "#,##0"

        c4 = ws.cell(row=current_row, column=4, value=pct)
        c4.alignment = Alignment(horizontal="right")
        c4.number_format = "0.0%"

        c5 = ws.cell(row=current_row, column=5, value=districts)
        c5.alignment = Alignment(horizontal="right")
        c5.number_format = "#,##0"

        for c in range(1, 6):
            ws.cell(row=current_row, column=c).font = DATA_FONT
            ws.cell(row=current_row, column=c).border = THIN_BORDER

        current_row += 1

    # Total Row
    ws.cell(row=current_row, column=1, value="")
    ws.cell(row=current_row, column=2, value="TOTAL ALL INDIA").font = BOLD_FONT
    tot_c3 = ws.cell(row=current_row, column=3, value=f"=SUM(C{header_row + 1}:C{current_row - 1})")
    tot_c3.font = BOLD_FONT
    tot_c3.alignment = Alignment(horizontal="right")
    tot_c3.number_format = "#,##0"

    tot_c4 = ws.cell(row=current_row, column=4, value=f"=SUM(D{header_row + 1}:D{current_row - 1})")
    tot_c4.font = BOLD_FONT
    tot_c4.alignment = Alignment(horizontal="right")
    tot_c4.number_format = "0.0%"

    tot_c5 = ws.cell(row=current_row, column=5, value=df["District / City"].nunique())
    tot_c5.font = BOLD_FONT
    tot_c5.alignment = Alignment(horizontal="right")
    tot_c5.number_format = "#,##0"

    for c in range(1, 6):
        ws.cell(row=current_row, column=c).border = TOP_THICK_BOTTOM_DOUBLE

    # Adjust summary column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 20

def export_db(db_path=DB_PATH, excel_path=EXCEL_PATH, csv_path=CSV_PATH):
    """Main export pipeline from SQLite to Excel and CSV."""
    if not os.path.exists(db_path):
        print(f"[!] Error: Database {db_path} not found!")
        return

    print("=" * 60)
    print(" EXPORTING NAYARA ENERGY DATABASE TO EXCEL & CSV")
    print("=" * 60)

    conn = sqlite3.connect(db_path)
    df = pd.read_sql_query("SELECT * FROM outlets ORDER BY state, district, ro_name", conn)
    conn.close()

    if df.empty:
        print("[!] Database is empty! No records to export.")
        return

    df = df.rename(columns=DISPLAY_COLUMNS)

    # 1. Export CSV
    print(f"  * Writing CSV: {csv_path} ({len(df):,} records)...")
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")

    # 2. Export Multi-sheet Excel
    print(f"  * Building Excel Workbook: {excel_path}...")
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="All Outlets", index=False)

        top_states = df["State / UT"].value_counts().head(5).index
        for state in top_states:
            if not state or pd.isna(state):
                continue
            state_df = df[df["State / UT"] == state]
            sheet_name = str(state)[:28]
            state_df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = openpyxl.load_workbook(excel_path)

    if "All Outlets" in wb.sheetnames:
        style_data_sheet(wb["All Outlets"], df, "All Outlets Master Database")

    for state in top_states:
        if not state or pd.isna(state):
            continue
        sheet_name = str(state)[:28]
        if sheet_name in wb.sheetnames:
            state_df = df[df["State / UT"] == state]
            style_data_sheet(wb[sheet_name], state_df, f"Stations in {state}")

    build_summary_sheet(wb, df)

    wb.save(excel_path)
    wb.close()

    print("=" * 60)
    print(" [OK] EXPORT COMPLETED SUCCESSFULLY")
    print(f"  * Master Excel:  {excel_path}")
    print(f"  * Master CSV:    {csv_path}")
    print(f"  * Total Records: {len(df):,}")
    print("=" * 60)

if __name__ == "__main__":
    export_db()
